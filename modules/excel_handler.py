"""
Module de gestion des fichiers Excel
Import et Export
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

class ExcelHandler:
    
    @staticmethod
    def lire_fichier_factures(chemin_fichier):
        """
        Lit le fichier Excel des factures (4 colonnes)
        Colonnes: design, SommeDeMontant, date, CODE MAGASIN
        """
        try:
            df = pd.read_excel(chemin_fichier)
            
            # Nettoyer les noms de colonnes
            df.columns = df.columns.str.strip()
            
            # Vérifier colonnes requises
            colonnes_requises = ['design', 'SommeDeMontant', 'date', 'CODE MAGASIN']
            
            # Mapper colonnes si noms différents
            mapping = {
                'design': 'design',
                'SommeDeMontant': 'montant',
                'date': 'date',
                'CODE MAGASIN': 'code'
            }
            
            # Renommer
            for old, new in mapping.items():
                if old in df.columns:
                    df.rename(columns={old: new}, inplace=True)
            
            # Convertir en liste de dict
            return df.to_dict('records')
            
        except Exception as e:
            raise Exception(f"Erreur lecture fichier factures: {str(e)}")
    
    @staticmethod
    def lire_fichier_activations(chemin_fichier):
        """
        Lit le fichier Excel des activations
        Colonnes importantes: N° Tel, Date Activation, Code POS, Plan Tarifaire, 
                             Indicateur VOIX/3G, commission d'activation nette
        """
        try:
            df = pd.read_excel(chemin_fichier)
            
            # Nettoyer colonnes
            df.columns = df.columns.str.strip()
            
            # Mapper colonnes
            activations = []
            
            for _, row in df.iterrows():
                act = {
                    'numero_tel': str(row.get('N° Tel', '')),
                    'date_activation': str(row.get('Date Activation', '')),
                    'code': str(row.get('Code POS', '')),
                    'nom_pos': str(row.get('Nom POS', '')),
                    'plan_tarifaire': str(row.get('Plan Tarifaire', '')),
                    'type_indication': str(row.get('Indicateur VOIX/3G', '')),
                    'montant': float(row.get('Montant forfait', 0)),
                    'commission': float(row.get('Commission d\'activation nette', 0) or 
                                      row.get('commission d\'activation nette', 0) or
                                      row.get('commission d\'activation nettte', 0) or 0),
                    'nom_client': str(row.get('Nom Titulaire du Client', '')),
                    'type_segment': str(row.get('Type de segment', '') or 
                                      row.get('Type segment', '') or 
                                      'Residential')  # Défaut si colonne absente
                }
                activations.append(act)
            
            return activations
            
        except Exception as e:
            raise Exception(f"Erreur lecture fichier activations: {str(e)}")
    
    @staticmethod
    def lire_base_magasins(chemin_fichier):
        """
        Lit le fichier Excel de la base magasins
        """
        try:
            df = pd.read_excel(chemin_fichier)
            
            # Nettoyer colonnes
            df.columns = df.columns.str.strip()
            
            magasins = []
            
            for _, row in df.iterrows():
                mag = {
                    'code': str(row.get('Code', '')),
                    'nom': str(row.get('Nom', '')),
                    'adresse': str(row.get('Adresse1', '') + ' ' + str(row.get('Ville', ''))).strip(),
                    'ville': str(row.get('Ville', '')),
                    'rc': str(row.get('RC', '')),
                    'if_num': str(row.get('IF', row.get('if', ''))),  # CORRECTION: IF en majuscule
                    'ice': str(row.get('ICE', '')),
                    'patente': str(row.get('Patente', '')),
                    'societe': str(row.get('Societe', 'BESTMARK')),
                    'type_pos': str(row.get('type pos', 'moral')).lower(),
                    'attestation': str(row.get('avec attestation', '')).lower().strip()
                }
                
                # Nettoyer attestation
                if mag['attestation'] in ['oui', 'yes', '1']:
                    mag['attestation'] = 'oui'
                elif mag['attestation'] in ['non', 'no', '0']:
                    mag['attestation'] = 'non'
                else:
                    mag['attestation'] = ''
                
                magasins.append(mag)
            
            return magasins
            
        except Exception as e:
            raise Exception(f"Erreur lecture base magasins: {str(e)}")
    
    @staticmethod
    def exporter_etat_excel(donnees_factures, chemin_sortie):
        """
        Exporte un état récapitulatif en Excel
        donnees_factures = liste de dict avec toutes les infos
        """
        try:
            # Créer DataFrame
            rows = []
            
            for fact in donnees_factures:
                row = {
                    'Code': fact['code'],
                    'Nom': fact['nom_magasin'],
                    'ville': fact.get('ville_magasin', ''),
                    'Date': fact.get('date_facture', ''),
                    'Société': fact['societe'],
                    'N° Facture': fact['numero_facture'],
                    'Mois': fact['mois_nom'],
                    'Année': fact['annee']
                }
                
                # Ajouter chaque produit (primes)
                for ligne in fact.get('lignes', []):
                    col_name = ligne['designation']
                    row[col_name] = ligne['montant']
                    
                    # Si c'est Commission inscription, ajouter colonne Qté
                    if 'inscription' in col_name.lower() and 'quantite' in ligne:
                        row['Qté Activations'] = ligne['quantite']
                
                # Totaux (APRÈS les primes)
                row['Total HT'] = fact['montant_ht']
                row['TVA 20%'] = fact['tva']
                row['Montant TTC'] = fact['ttc']
                
                if fact.get('ras_applicable'):
                    row['RAS 10%'] = -fact['ras']
                    row['TVA/RAS'] = -fact['tva_ras']
                
                row['NET À PAYER'] = fact['net_a_payer']
                
                rows.append(row)
            
            df = pd.DataFrame(rows)
            
            # Remplir cellules vides avec 0 (pour toutes les colonnes de primes)
            colonnes_info = ['Code', 'Nom', 'Société', 'N° Facture', 'Mois', 'Année']
            for col in df.columns:
                if col not in colonnes_info:
                    df[col] = df[col].fillna(0)
            
            # ORDRE EXACT DES COLONNES (selon fichier exemple)
            colonnes_ordre_fixe = [
                'Code', 'Nom', 'ville',
                'Date',
                'airtime prepaye',
                'Commission d\'inscription Mobile', 'Qté Activations',
                'Prime d\'accès Fixe',
                'Airtime postpaye',
                'Prime Permanence p1',
                'Prime Permanence P2 12',
                'Prime Permanence P2 24',
                'Prime d\'enseigne Mobile',
                'Prime d\'enseigne  FIXE',
                'Prime TPE',
                'Comm d\'activations Fixe',  # AVANT totaux
                'COMM fixe prestataire',
                'Prime d\'accès global',
                'Support Commercial',
                'Prime portabilité',
                'Prime d\'objectif Mobile',
                'Prime d\'objectif Fixe',
                'Kiosque',
                'pénalite resiliation',
                'penalite retard doc',
                'Total HT', 'TVA 20%', 'Montant TTC',
                'RAS 10%', 'TVA/RAS', 'NET À PAYER'
            ]
            
            # Garder seulement les colonnes qui existent
            colonnes_finales = [col for col in colonnes_ordre_fixe if col in df.columns]
            
            # Ajouter colonnes manquantes qui existent dans df mais pas dans l'ordre
            for col in df.columns:
                if col not in colonnes_finales and col not in ['Société', 'N° Facture', 'Mois', 'Année']:
                    colonnes_finales.append(col)
            
            df = df[colonnes_finales]
            
            # Créer Excel avec style
            with pd.ExcelWriter(chemin_sortie, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='État Factures', index=False)
                
                # Styler
                workbook = writer.book
                worksheet = writer.sheets['État Factures']
                
                # Header style
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Colorer colonnes totaux en jaune clair
                col_totaux_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                
                # Colonnes totaux à colorer
                colonnes_totaux = ['Total HT', 'TVA 20%', 'Montant TTC', 'RAS 10%', 'TVA/RAS', 'NET À PAYER']
                
                # Trouver indices colonnes totaux
                header_row = [cell.value for cell in worksheet[1]]
                indices_totaux = [i+1 for i, col in enumerate(header_row) if col in colonnes_totaux]
                
                # Appliquer couleur
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for idx in indices_totaux:
                        if idx <= len(row):
                            row[idx-1].fill = col_totaux_fill
                
                # Auto-ajuster colonnes
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            return True
            
        except Exception as e:
            raise Exception(f"Erreur export Excel: {str(e)}")
    
    @staticmethod
    def exporter_base_magasins(magasins_list, chemin_sortie):
        """Exporte la base magasins en Excel"""
        try:
            df = pd.DataFrame(magasins_list)
            
            # Réorganiser colonnes
            colonnes_ordre = ['code', 'nom', 'adresse', 'ville', 'rc', 'if', 'ice', 
                            'patente', 'societe', 'type_pos', 'attestation']
            
            df = df[[col for col in colonnes_ordre if col in df.columns]]
            
            df.to_excel(chemin_sortie, index=False, sheet_name='Base Magasins')
            
            return True
            
        except Exception as e:
            raise Exception(f"Erreur export base: {str(e)}")
    
    @staticmethod
    def valider_fichier_factures(donnees):
        """Valide que le fichier factures a la bonne structure"""
        if not donnees or len(donnees) == 0:
            return False, "Fichier vide"
        
        # Vérifier colonnes
        first = donnees[0]
        colonnes_requises = ['design', 'montant', 'date', 'code']
        
        for col in colonnes_requises:
            if col not in first:
                return False, f"Colonne manquante: {col}"
        
        return True, "OK"
    
    @staticmethod
    def valider_fichier_activations(donnees):
        """Valide que le fichier activations a la bonne structure"""
        if not donnees or len(donnees) == 0:
            return False, "Fichier vide"
        
        first = donnees[0]
        colonnes_requises = ['numero_tel', 'code', 'plan_tarifaire', 'commission']
        
        for col in colonnes_requises:
            if col not in first:
                return False, f"Colonne manquante: {col}"
        
        return True, "OK"