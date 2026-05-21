"""
create_fr_accounts.py — Crée les comptes FR pour tous les magasins.

Usage :
    python create_fr_accounts.py
    python create_fr_accounts.py --db /chemin/database_web.db --out comptes_fr.xlsx

Règles :
  - username  = "pos-{code}"
  - password  = 8 caractères aléatoires (lettres + chiffres)
  - role      = "fr"
  - N'écrase pas les comptes déjà existants
  - Exporte un fichier Excel avec tous les comptes (nouveaux + existants)
"""

import sqlite3
import os
import random
import string
import argparse
from datetime import datetime

from werkzeug.security import generate_password_hash

BASE_DIR     = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
DEFAULT_DB   = os.environ.get(
    'DATABASE_PATH',
    os.path.join(os.path.dirname(__file__), 'config', 'database_web.db')
)
DEFAULT_OUT  = os.path.join(os.path.dirname(__file__), 'comptes_fr.xlsx')


def gen_password(length=8):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choices(chars, k=length))


def ensure_column(con, table, column, col_type='INTEGER', default='0'):
    """Ajoute une colonne si elle n'existe pas encore."""
    cols = [r[1] for r in con.execute(f'PRAGMA table_info({table})').fetchall()]
    if column not in cols:
        con.execute(f'ALTER TABLE {table} ADD COLUMN {column} {col_type} NOT NULL DEFAULT {default}')
        con.commit()
        print(f'  [DB] Colonne "{column}" ajoutée à {table}.')


def run(db_path=DEFAULT_DB, out_path=DEFAULT_OUT, reset_passwords=False):
    if not os.path.exists(db_path):
        raise FileNotFoundError(f'Base de données introuvable : {db_path}')

    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row

    # S'assurer que must_change_password existe
    ensure_column(con, 'web_users', 'must_change_password', 'INTEGER', '0')

    # Charger tous les magasins
    magasins = con.execute(
        'SELECT code, nom, societe FROM magasins ORDER BY societe, code'
    ).fetchall()

    print(f'\n[create_fr_accounts] {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print(f'  DB      : {db_path}')
    print(f'  Export  : {out_path}')
    print(f'  Magasins: {len(magasins)}\n')

    mode = 'RESET PASSWORDS' if reset_passwords else 'CRÉATION'
    print(f'  Mode    : {mode}\n')

    rapport       = []   # [(code, nom, societe, username, password, statut)]
    crees         = 0
    resets        = 0
    ignores       = 0

    for mag in magasins:
        code     = str(mag['code'])
        nom      = mag['nom'] or ''
        societe  = mag['societe'] or ''
        username = f'pos-{code}'

        existing = con.execute(
            'SELECT id FROM web_users WHERE username = ?', (username,)
        ).fetchone()

        if existing and reset_passwords:
            password = gen_password()
            ph = generate_password_hash(password)
            con.execute(
                '''UPDATE web_users
                   SET password_hash = ?, must_change_password = 1
                   WHERE username = ?''',
                (ph, username)
            )
            rapport.append((code, nom, societe, username, password, 'reset'))
            resets += 1

        elif existing:
            rapport.append((code, nom, societe, username, '(déjà existant)', 'existant'))
            ignores += 1

        else:
            password = gen_password()
            ph = generate_password_hash(password)
            now = datetime.utcnow().isoformat()
            con.execute(
                '''INSERT INTO web_users
                   (username, email, password_hash, role, societe, code_magasin,
                    actif, must_change_password, created_at)
                   VALUES (?, NULL, ?, 'fr', ?, ?, 1, 1, ?)''',
                (username, ph, societe, code, now)
            )
            rapport.append((code, nom, societe, username, password, 'créé'))
            crees += 1

    con.commit()
    con.close()

    if reset_passwords:
        print(f'  Réinitialisés : {resets}')
        print(f'  Nouveaux      : {crees}')
    else:
        print(f'  Créés   : {crees}')
        print(f'  Ignorés : {ignores} (comptes déjà existants)')

    # ── Export Excel ────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print('\n  [ERREUR] openpyxl non installé. Lancez : pip install openpyxl')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comptes FR'

    # En-têtes
    headers = ['Code', 'Nom Magasin', 'Société', 'Username', 'Mot de passe', 'Statut']
    header_fill   = PatternFill('solid', fgColor='1F4E79')
    header_font   = Font(bold=True, color='FFFFFF', size=11)
    header_align  = Alignment(horizontal='center', vertical='center')
    thin          = Side(style='thin', color='CCCCCC')
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 22
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border

    # Largeurs colonnes
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12

    # Couleurs par statut / société
    fill_best  = PatternFill('solid', fgColor='DBEAFE')   # bleu clair  — BESTMARK
    fill_afri  = PatternFill('solid', fgColor='D1FAE5')   # vert clair  — AFRINETWORKS
    fill_mdp   = PatternFill('solid', fgColor='FEF9C3')   # jaune       — cellule mot de passe
    fill_exist = PatternFill('solid', fgColor='F3F4F6')   # gris        — non modifié
    fill_reset = PatternFill('solid', fgColor='FEE2E2')   # rouge clair — mot de passe réinitialisé

    for row_idx, (code, nom, societe, username, password, statut) in enumerate(rapport, 2):
        ws.row_dimensions[row_idx].height = 18
        values = [code, nom, societe, username, password, statut]

        if statut == 'existant':
            row_fill = fill_exist
        elif societe == 'BESTMARK':
            row_fill = fill_best
        else:
            row_fill = fill_afri

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border    = border
            cell.alignment = Alignment(vertical='center')
            if col == 5 and statut in ('créé', 'reset'):
                cell.fill = fill_reset if statut == 'reset' else fill_mdp
            else:
                cell.fill = row_fill

    # Geler la ligne d'en-tête + filtre auto
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)
    print(f'\n  Export Excel : {out_path}')
    print(f'  {len(rapport)} lignes exportées.\n')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Crée les comptes FR POS')
    parser.add_argument('--db',  default=DEFAULT_DB,  help='Chemin database_web.db')
    parser.add_argument('--out', default=DEFAULT_OUT, help='Chemin fichier Excel de sortie')
    parser.add_argument('--reset-passwords', action='store_true',
                        help='Régénère les mots de passe de tous les comptes existants')
    args = parser.parse_args()

    try:
        run(db_path=args.db, out_path=args.out, reset_passwords=args.reset_passwords)
    except FileNotFoundError as e:
        print(f'[ERREUR] {e}')
        raise SystemExit(1)
