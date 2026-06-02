import io
import json
import os
import sqlite3
import sys
from datetime import datetime

from flask import (Blueprint, render_template, request,
                   abort, send_file, flash, redirect, url_for)
from flask_login import login_required, current_user

from utils import role_required

liquidation_bp = Blueprint('liquidation', __name__)

MOIS_NOMS = [
    'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
    'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre',
]

_PROJECT_ROOT  = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
_WEB_ROOT      = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

DB_PATH = os.environ.get(
    'DATABASE_PATH',
    os.path.join(os.path.dirname(__file__), '..', 'config', 'database_web.db')
)

# (colonne DB, libellé affiché)
COLONNES = [
    ('airtime_prepaye',   'Airtime Prépayé'),
    ('airtime_postpaye',  'Airtime Postpayé'),
    ('prime_inscription', 'Prime Inscription'),
    ('prime_acces',       "Prime d'Accès"),
    ('enseigne_mobile',   'Enseigne Mobile'),
    ('enseigne_fixe',     'Enseigne Fixe'),
    ('objectif_mobile',   'Obj. Mobile'),
    ('objectif_fixe',     'Obj. Fixe'),
    ('penalite',          'Pénalité'),
    ('tpe',               'TPE'),
]

# colonne DB → clé retournée par LiquidationProcessor
_PROC_KEY = {
    'airtime_prepaye':   'airtime_prepaye',
    'airtime_postpaye':  'airtime_postpaye',
    'prime_inscription': 'prime_inscription',
    'prime_acces':       'prime_acces',
    'enseigne_mobile':   'prime_enseigne_mobile',
    'enseigne_fixe':     'prime_enseigne_fixe',
    'objectif_mobile':   'prime_objectif_mobile',
    'objectif_fixe':     'prime_objectif_fixe',
    'penalite':          'penalite',
    'tpe':               'prime_tpe',
}

CHAMPS_DB = [col for col, _ in COLONNES]


# ── Helpers DB ───────────────────────────────────────────────────────────────

def _db():
    con = sqlite3.connect(os.path.abspath(DB_PATH))
    con.row_factory = sqlite3.Row
    return con


def init_tables():
    con = sqlite3.connect(os.path.abspath(DB_PATH))
    con.execute('''
        CREATE TABLE IF NOT EXISTS liquidations_web (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            mois              INTEGER NOT NULL,
            annee             INTEGER NOT NULL,
            societe           TEXT    NOT NULL,
            code              TEXT    NOT NULL,
            nom               TEXT,
            airtime_prepaye   REAL DEFAULT 0,
            airtime_postpaye  REAL DEFAULT 0,
            prime_inscription REAL DEFAULT 0,
            prime_acces       REAL DEFAULT 0,
            enseigne_mobile   REAL DEFAULT 0,
            enseigne_fixe     REAL DEFAULT 0,
            objectif_mobile   REAL DEFAULT 0,
            objectif_fixe     REAL DEFAULT 0,
            penalite          REAL DEFAULT 0,
            tpe               REAL DEFAULT 0,
            total             REAL DEFAULT 0,
            created_at        TEXT,
            created_by        TEXT,
            UNIQUE(mois, annee, societe, code)
        )
    ''')
    con.commit()
    con.close()


# ── Adaptateur DB pour LiquidationProcessor ─────────────────────────────────

class _DBAdapter:
    def __init__(self, path):
        self._path = path

    def get_connection(self):
        return sqlite3.connect(self._path)


# ── Traitement ───────────────────────────────────────────────────────────────

def _charger_processor():
    if _WEB_ROOT not in sys.path:
        sys.path.insert(0, _WEB_ROOT)
    from modules.liquidation_processor import LiquidationProcessor
    proc = LiquidationProcessor()
    coef_path = os.path.join(_PROJECT_ROOT, 'config', 'coefficients.json')
    if os.path.exists(coef_path):
        with open(coef_path, 'r', encoding='utf-8') as f:
            proc.coefficients = json.load(f)
    return proc


def _traiter_fichiers(fichiers, societe):
    """fichiers : dict {champ: BytesIO}. Retourne {code: {proc_key: valeur}}."""
    proc        = _charger_processor()
    mobile_data = proc.lire_fichier_mobile(fichiers['mobile'])      if fichiers.get('mobile')    else {}
    darbox_data = proc.lire_fichier_darbox(fichiers['darbox'])      if fichiers.get('darbox')    else {}
    b2b_data    = proc.lire_fichier_fixe_b2b(fichiers['fixe_b2b']) if fichiers.get('fixe_b2b') else {}
    b2c_data    = proc.lire_fichier_fixe_b2c(fichiers['fixe_b2c']) if fichiers.get('fixe_b2c') else {}
    resultats = proc.calculer_etat_complet(
        mobile_data, darbox_data, b2b_data, b2c_data,
        _DBAdapter(os.path.abspath(DB_PATH)),
        societe,
    )
    if resultats:
        premier_code, premier_vals = next(iter(resultats.items()))
        print("DEBUG résultats processor:", json.dumps(
            {premier_code: {k: v for k, v in premier_vals.items() if not k.startswith('_')}},
            ensure_ascii=False, default=str
        ))
    return resultats


# ── Routes ───────────────────────────────────────────────────────────────────

@liquidation_bp.route('/')
@login_required
@role_required('admin', 'agent')
def index():
    f_mois    = request.args.get('mois',    type=int)
    f_annee   = request.args.get('annee',   type=int)
    f_societe = request.args.get('societe', '').strip()

    con = _db()

    # Fichiers importés depuis le desktop (présence des 4 fichiers source)
    sql    = 'SELECT * FROM fichiers_liquidation WHERE 1=1'
    params = []
    if f_mois:
        sql += ' AND mois = ?';    params.append(f_mois)
    if f_annee:
        sql += ' AND annee = ?';   params.append(f_annee)
    if f_societe:
        sql += ' AND societe = ?'; params.append(f_societe)
    sql += ' ORDER BY annee DESC, mois DESC, societe'

    try:
        fl_rows = con.execute(sql, params).fetchall()
    except sqlite3.OperationalError:
        fl_rows = []

    # Résultats calculés (agrégés depuis liquidations_web)
    try:
        liq_rows = con.execute(
            'SELECT mois, annee, societe, COUNT(*) as nb_codes, SUM(total) as total_commissions '
            'FROM liquidations_web GROUP BY mois, annee, societe'
        ).fetchall()
        liq_index = {(r['mois'], r['annee'], r['societe']): dict(r) for r in liq_rows}
    except sqlite3.OperationalError:
        liq_index = {}

    con.close()

    liquidations = []
    for r in fl_rows:
        d = dict(r)
        m = d.get('mois')
        d['mois_nom']   = MOIS_NOMS[m - 1] if m and 1 <= m <= 12 else str(m or '')
        d['a_mobile']   = bool(d.get('fichier_mobile'))
        d['a_darbox']   = bool(d.get('fichier_darbox'))
        d['a_fixe_b2b'] = bool(d.get('fichier_fixe_b2b'))
        d['a_fixe_b2c'] = bool(d.get('fichier_fixe_b2c'))
        d['calcule']    = liq_index.get((d['mois'], d['annee'], d['societe']))
        liquidations.append(d)

    return render_template('liquidation/index.html',
                           liquidations=liquidations,
                           MOIS_NOMS=MOIS_NOMS,
                           f_mois=f_mois, f_annee=f_annee, f_societe=f_societe,
                           now=datetime.now())


@liquidation_bp.route('/import', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'agent')
def import_liq():
    pre_mois    = request.args.get('mois',    type=int)
    pre_annee   = request.args.get('annee',   type=int)
    pre_societe = request.args.get('societe', '')

    if request.method == 'POST':
        mois    = request.form.get('mois',    type=int)
        annee   = request.form.get('annee',   type=int)
        societe = request.form.get('societe', '').strip()

        def _render_form(**kw):
            return render_template('liquidation/import.html',
                                   MOIS_NOMS=MOIS_NOMS, now=datetime.now(),
                                   pre_mois=mois, pre_annee=annee, pre_societe=societe,
                                   **kw)

        if not mois or not annee or not societe:
            flash('Mois, année et société sont obligatoires.', 'danger')
            return _render_form()

        # Lire les fichiers en mémoire — aucune écriture sur disque
        fichiers = {}
        for champ in ('mobile', 'darbox', 'fixe_b2b', 'fixe_b2c'):
            f = request.files.get(champ)
            if f and f.filename:
                fichiers[champ] = io.BytesIO(f.read())

        if not fichiers:
            flash('Aucun fichier uploadé.', 'danger')
            return _render_form()

        for buf in fichiers.values():
            buf.seek(0)

        # Calcul des primes via LiquidationProcessor
        try:
            etat = _traiter_fichiers(fichiers, societe)
        except ImportError:
            flash('Module liquidation_processor introuvable. '
                  'Vérifiez que le dossier modules/ est accessible.', 'danger')
            return _render_form()
        except Exception as e:
            flash(f'Erreur traitement : {e}', 'danger')
            return _render_form()

        # Noms des magasins pour la colonne nom
        con      = _db()
        mag_noms = {str(r['code']): r['nom']
                    for r in con.execute('SELECT code, nom FROM magasins').fetchall()}

        # Supprimer les anciens résultats pour ce mois/annee/societe
        con.execute('DELETE FROM liquidations_web WHERE mois=? AND annee=? AND societe=?',
                    (mois, annee, societe))

        now_str    = datetime.now().isoformat(timespec='seconds')
        nb_codes   = 0
        total_glob = 0.0

        for code, primes in etat.items():
            vals  = {col: round(float(primes.get(_PROC_KEY[col]) or 0), 4) for col in CHAMPS_DB}
            total = sum(vals.values())
            total_glob += total
            nb_codes   += 1
            con.execute('''
                INSERT OR REPLACE INTO liquidations_web
                    (mois, annee, societe, code, nom,
                     airtime_prepaye, airtime_postpaye, prime_inscription, prime_acces,
                     enseigne_mobile, enseigne_fixe, objectif_mobile, objectif_fixe,
                     penalite, tpe, total, created_at, created_by)
                VALUES (?,?,?,?,?, ?,?,?,?, ?,?,?,?, ?,?,?, ?,?)
            ''', (
                mois, annee, societe, str(code), mag_noms.get(str(code), ''),
                vals['airtime_prepaye'],   vals['airtime_postpaye'],
                vals['prime_inscription'], vals['prime_acces'],
                vals['enseigne_mobile'],   vals['enseigne_fixe'],
                vals['objectif_mobile'],   vals['objectif_fixe'],
                vals['penalite'],          vals['tpe'],
                round(total, 4),
                now_str, current_user.username,
            ))

        con.commit()
        con.close()

        flash(
            f'Liquidation {MOIS_NOMS[mois - 1]} {annee} — {societe} calculée : '
            f'{nb_codes} codes, total {total_glob:,.2f} MAD.',
            'success',
        )
        return redirect(url_for('liquidation.detail', mois=mois, annee=annee, societe=societe))

    return render_template('liquidation/import.html',
                           MOIS_NOMS=MOIS_NOMS, now=datetime.now(),
                           pre_mois=pre_mois, pre_annee=pre_annee, pre_societe=pre_societe)


@liquidation_bp.route('/<int:mois>/<int:annee>/<societe>/detail')
@login_required
@role_required('admin', 'agent')
def detail(mois, annee, societe):
    con      = _db()
    liq_rows = con.execute(
        'SELECT * FROM liquidations_web WHERE mois=? AND annee=? AND societe=? ORDER BY code',
        (mois, annee, societe)
    ).fetchall()
    con.close()

    if not liq_rows:
        abort(404)

    mois_nom = MOIS_NOMS[mois - 1] if 1 <= mois <= 12 else str(mois)

    lignes = []
    for r in liq_rows:
        r = dict(r)
        lignes.append({
            'code':   r['code'],
            'nom':    r['nom'] or '—',
            'primes': {col: r.get(col, 0) for col, _ in COLONNES},
            'total':  r['total'],
        })

    col_totals   = {col: sum(float(l['primes'].get(col) or 0) for l in lignes)
                    for col, _ in COLONNES}
    total_global = sum(l['total'] for l in lignes)

    row = {
        'mois':       mois,
        'annee':      annee,
        'societe':    societe,
        'nb_codes':   len(lignes),
        'created_at': liq_rows[0]['created_at'],
        'created_by': liq_rows[0]['created_by'],
    }

    return render_template('liquidation/detail.html',
                           row=row, mois_nom=mois_nom,
                           lignes=lignes, colonnes=COLONNES,
                           col_totals=col_totals, total_global=total_global)


@liquidation_bp.route('/<int:mois>/<int:annee>/<societe>/export')
@login_required
@role_required('admin', 'agent')
def export(mois, annee, societe):
    con      = _db()
    liq_rows = con.execute(
        'SELECT * FROM liquidations_web WHERE mois=? AND annee=? AND societe=? ORDER BY code',
        (mois, annee, societe)
    ).fetchall()
    con.close()

    if not liq_rows:
        abort(404)

    mois_nom = MOIS_NOMS[mois - 1] if 1 <= mois <= 12 else str(mois)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{mois_nom[:3]} {annee}"

    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    num_fmt  = '#,##0.00'

    headers = ['Code', 'Nom Magasin'] + [lbl for _, lbl in COLONNES] + ['TOTAL']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    total_global = 0.0
    col_totals   = {col: 0.0 for col, _ in COLONNES}

    for ri, r in enumerate(liq_rows, 2):
        r     = dict(r)
        total = r.get('total', 0)
        total_global += total
        vals = [r['code'], r['nom'] or '']
        for col, _ in COLONNES:
            v = round(float(r.get(col) or 0), 2)
            col_totals[col] += v
            vals.append(v)
        vals.append(round(total, 2))
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if ci > 2:
                cell.number_format = num_fmt

    last     = len(liq_rows) + 2
    tot_fill = PatternFill('solid', fgColor='EFF6FF')
    tot_font = Font(bold=True, color='1E40AF', size=10)
    ws.cell(row=last, column=2, value='TOTAL').font = tot_font
    for ci, (col, _) in enumerate(COLONNES, 3):
        cell = ws.cell(row=last, column=ci, value=round(col_totals[col], 2))
        cell.font = tot_font; cell.fill = tot_fill; cell.number_format = num_fmt
    grand = ws.cell(row=last, column=len(headers), value=round(total_global, 2))
    grand.font = tot_font; grand.fill = tot_fill; grand.number_format = num_fmt

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 24
    for ci in range(3, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"liquidation_{societe}_{annee}_{mois:02d}.xlsx",
    )
